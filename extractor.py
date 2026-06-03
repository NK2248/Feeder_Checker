#!/usr/bin/env python3
# pip install openpyxl
"""
IO Point Extractor - 115 kV SCPS H Scheme
Output: 3 CSV files per source sheet (Status / Control / Analog).
Requires: pip install openpyxl
"""
import sys, csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed.  Run:  pip install openpyxl")
    input("Press Enter to exit...")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
SOURCE_FILE = r"C:\CU\EIC\Intern\Feeder_Checker\OG\Substation\[Standard-2025] IO Point Lists for 115 kV SCPS H Scheme.xlsx"
OUTPUT_DIR  = r"C:\CU\EIC\Intern\Feeder_Checker\Substation_csv"
# ─────────────────────────────────────────────────────────────────────────────

SKIP_SHEETS = {"COVER", "PROJ_DESC", "IP_ADDRESS"}

STATUS_COLS  = ["ITEM","FEEDER_NAME","BAY_NAME","SPECIFICATION_POINT_NAME","SPECIFICATION_DESCRIPTOR",
                "STATE_0","STATE_1","STATE_2","STATE_3",
                "DNP3.0_OBJ","DNP3.0_VAR","DNP3.0_QII","DNP3.0_CLASS","DNP3.0_TYPE","DNP3.0_ADDRESS",
                "REMARK","NOTE","BAY_PROT.&BCU_NAME","IEC61850_PROT.&BCU_REFERENCE","BRCB/URCB"]
CONTROL_COLS = ["ITEM","FEEDER_NAME","BAY_NAME","SPECIFICATION_POINT_NAME","SPECIFICATION_DESCRIPTOR",
                "STATE_CLOSE","STATE_TRIP",
                "DNP3.0_OBJ","DNP3.0_VAR","DNP3.0_QII","DNP3.0_CLASS","DNP3.0_TYPE","DNP3.0_ADDRESS",
                "REMARK","NOTE","BAY_PROT.&BCU_NAME","IEC61850_PROT.&BCU_REFERENCE","BRCB/URCB"]
ANALOG_COLS  = ["ITEM","FEEDER_NAME","BAY_NAME","SPECIFICATION_POINT_NAME","SPECIFICATION_DESCRIPTOR",
                "SPECIFICATION_UNIT","SCALE_ACTUAL_DATA","SCALE_RAW_DATA",
                "DNP3.0_OBJ","DN3.0_VAR","DNP3.0_QII","DNP3.0_CLASS","DNP3.0_TYPE","DNP3.0_ADDRESS",
                "REMARK","NOTE","BAY_PROT.&BCU_NAME","IEC61850_PROT.&BCU_REFERENCE","BRCB/URCB"]

def clean(v): return "" if v is None else str(v).strip()
def is_section(v, kw): return bool(v and isinstance(v, str) and kw.upper() in v.upper())
def get(row, idx):
    try: return clean(row[idx])
    except: return ""

def extract_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    status, control, analog = [], [], []
    section = None
    for row in rows:
        first = row[0]
        if is_section(first, "STATUS POINT"):   section = "status";  continue
        if is_section(first, "CONTROL OUTPUT"): section = "control"; continue
        if is_section(first, "ANALOG POINT"):   section = "analog";  continue
        if is_section(first, "ALARM POINT"):    section = None;      continue
        if section is None: continue
        if not isinstance(first, (int, float)): continue
        r = row
        if section == "status":
            status.append([get(r,0),get(r,1),get(r,2),get(r,3),get(r,4),
                           get(r,5),get(r,6),get(r,7),get(r,8),
                           get(r,9),get(r,10),get(r,11),get(r,12),get(r,13),get(r,14),
                           get(r,15),get(r,17),get(r,25),get(r,26),get(r,27)])
        elif section == "control":
            control.append([get(r,0),get(r,1),get(r,2),get(r,3),get(r,4),
                            get(r,5),get(r,7),
                            get(r,9),get(r,10),get(r,11),get(r,12),get(r,13),get(r,14),
                            get(r,15),get(r,17),get(r,25),get(r,26),get(r,27)])
        elif section == "analog":
            analog.append([get(r,0),get(r,1),get(r,2),get(r,3),get(r,4),
                           get(r,5),get(r,6),get(r,7),
                           get(r,9),get(r,10),get(r,11),get(r,12),get(r,13),get(r,14),
                           get(r,15),get(r,17),get(r,25),get(r,26),get(r,27)])
    return status, control, analog

def write_csv(path, headers, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    print(f"    Saved: {path.name}  ({len(rows)} rows)")

def slugify(name):
    result = ""
    for ch in name:
        result += ch if (ch.isalnum() or ch in "-_") else "_"
    return result.strip("_")

def main():
    src = Path(SOURCE_FILE)
    out = Path(OUTPUT_DIR)

    if not src.exists():
        print(f"ERROR: File not found: {src}")
        print("Edit SOURCE_FILE at the top of this script.")
        input("Press Enter to exit...")
        return

    out.mkdir(parents=True, exist_ok=True)
    print(f"Loading: {src.name}\n")

    wb = load_workbook(src, data_only=True)
    files_created = 0

    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue

        ws = wb[name]
        slug = slugify(name)
        st, ct, an = extract_sheet(ws)

        print(f"  [{name}]  {len(st)} status  |  {len(ct)} control  |  {len(an)} analog")

        if st:
            write_csv(out / f"{slug}_Status_point.csv",   STATUS_COLS,  st)
            files_created += 1
        if ct:
            write_csv(out / f"{slug}_Control_output.csv", CONTROL_COLS, ct)
            files_created += 1
        if an:
            write_csv(out / f"{slug}_Analog_point.csv",   ANALOG_COLS,  an)
            files_created += 1

    print(f"\nDone! {files_created} CSV files saved to:\n{out}")
    input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()